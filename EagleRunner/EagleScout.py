import glob
import os
import sqlite3
import openpyxl
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

TEAM = 254

os.chdir(r"C:/Users/Logan's Computer/PythonCode/EagleScout_Practice")

if os.path.exists("EagleRunner/Spreadsheets/combined.csv"):
    os.remove("EagleRunner/Spreadsheets/combined.csv")

if os.path.exists("EagleRunner/Spreadsheets/Combined_Raw.db"):
    os.remove("EagleRunner/Spreadsheets/Combined_Raw.db")

if os.path.exists("Tournaments/All_Tournaments2021.xlsx"):
    os.remove("Tournaments/All_Tournaments2021.xlsx")

if os.path.exists("EagleRunner/Spreadsheets/Combined.xlsx"):
    os.remove("EagleRunner/Spreadsheets/Combined.xlsx")

if os.path.exists("EagleRunner/Spreadsheets/ScoreContrib.xlsx"):
    os.remove("EagleRunner/Spreadsheets/ScoreContrib.xlsx")

if os.path.exists("EagleRunner/Spreadsheets/Master.xlsx"):
    os.remove("EagleRunner/Spreadsheets/Master.xlsx")

if os.path.exists("EagleRunner/Spreadsheets/Temp.xlsx"):
    os.remove("EagleRunner/Spreadsheets/Temp.xlsx")

if os.path.exists("Match_Schedule/Match Schedule.csv"):
    os.remove("Match_Schedule/Match Schedule.csv")

con = sqlite3.connect("EagleRunner/Spreadsheets/Combined_Raw.db")

redFill = PatternFill(start_color='F95555', end_color='F95555', fill_type='solid')
yellowFill = PatternFill(start_color='FBFE46', end_color='FBFE46', fill_type='solid')
greenFill = PatternFill(start_color='ACF99D', end_color='ACF99D', fill_type='solid')
clearFill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

files_in_dir = [f[10:] for f in glob.glob('Data_CSVs\*.csv')]

StrToInt_dict = {'Team': int, 'Match_Num': int, 'Auto_Cross': int, 'Auto_Outter': int, 'Auto_Bottom': int,
                 'Tele Outter': int, 'Tele_Bottom': int, 'Rotation': int, 'Position': int, 'Climb': int, 'Level': int,
                 'Driver_Perf': int, 'Auto_Perf': int, 'Name': str, 'Comments': str, 'Point_Contrib': int}

d1 = pd.read_csv("EagleRunner/Spreadsheets/Configurations/Header.txt")
d1.to_csv("EagleRunner/Spreadsheets/combined.csv", header=True, index=False)

fout = open("EagleRunner/Spreadsheets/combined.csv", "a")

for filenames in files_in_dir:
    fName, fExt = (os.path.splitext(filenames))
    sName = fName.split('-')
    N = (sName[1])
    T = (sName[0])

    for line in open("Data_CSVs/" + filenames):
        fout.write(str(T) + ",")
        fout.write(str(N) + ",")
        fout.write(line)
fout.close()

with pd.ExcelWriter("EagleRunner/Spreadsheets/Combined.xlsx",
                    engine='xlsxwriter') as writer:
    dt = pd.read_csv("EagleRunner/Spreadsheets/combined.csv")
    rows = len(dt.index)

    dt.to_excel(writer, sheet_name='All data', index=False)
    worksheet = writer.sheets['All data']
    writer.save()

ScrCont = openpyxl.load_workbook("EagleRunner/Spreadsheets/Combined.xlsx", read_only=False,
                                 keep_vba=True)

SN = ScrCont['All data']

ROWS = SN.max_row

for i in range(2, ROWS + 1):
    SC = SN.cell(row=i, column=3).value * 5
    SC += SN.cell(row=i, column=4).value * 4
    SC += SN.cell(row=i, column=5).value * 2
    SC += SN.cell(row=i, column=6).value * 2
    SC += SN.cell(row=i, column=7).value * 1
    SC += SN.cell(row=i, column=8).value * 10
    SC += SN.cell(row=i, column=9).value * 20
    if SN.cell(row=i, column=10).value == 0:
        val = 0
    elif SN.cell(row=i, column=10).value == 1:
        val = 5
    elif SN.cell(row=i, column=10).value == 2:
        val = 25
    elif SN.cell(row=i, column=10).value == 3:
        val = 50
    elif SN.cell(row=i, column=10).value == 4:
        val = 75

    SC += val
    SC += SN.cell(row=i, column=11).value * 15
    SN.cell(row=i, column=16).value = SC

ScrCont.save("EagleRunner/Spreadsheets/ScoreContrib.xlsx")

db = pd.read_excel("EagleRunner/Spreadsheets/ScoreContrib.xlsx")
db.to_sql("Raw_Data", con, if_exists='replace', index=False)

with pd.ExcelWriter("EagleRunner/Spreadsheets/Master.xlsx") as writer:
    df2 = pd.read_excel("EagleRunner/Spreadsheets/ScoreContrib.xlsx", converters=StrToInt_dict)

    group = df2.groupby("Team")
    for Team, Team_df in group:
        Team_df.to_excel(writer, sheet_name=("T" + str(Team)),
                         index=False)

    writer.save()

Data = pd.ExcelFile("EagleRunner/Spreadsheets/Master.xlsx")
Teams = Data.sheet_names

Tnmt = openpyxl.load_workbook("EagleRunner/Spreadsheets/Master.xlsx", read_only=False, keep_vba=True)

WS1 = Tnmt.create_sheet("Important Stuff", 0)
WS2 = Tnmt.create_sheet("Predictions")

for sht in Teams:
    sn = Tnmt[sht]
    sn['B16'] = str('Average')
    sn['B17'] = str('Standard Deviation')
    sn['B18'] = str('STDev % of Average')
    sn['B19'] = str('Total')
    sn['E21'] = str('Upper Ave')
    sn['E22'] = str('Lower Ave')
    sn['E23'] = str('Rotation Ave')
    sn['E24'] = str('Position Ave')
    sn['E25'] = str('Climb Ave')
    sn['E26'] = str('Points Ave')

    sn.cell(row=16, column=3).value = "=AVERAGE(C2:C13)"
    sn.cell(row=16, column=4).value = "=AVERAGE(D2:D13)"
    sn.cell(row=16, column=5).value = "=AVERAGE(E2:E13)"
    sn.cell(row=16, column=6).value = "=AVERAGE(F2:F13)"
    sn.cell(row=16, column=7).value = "=AVERAGE(G2:G13)"
    sn.cell(row=16, column=8).value = "=AVERAGE(H2:H13)"
    sn.cell(row=16, column=9).value = "=AVERAGE(I2:I13)"
    sn.cell(row=16, column=10).value = "=AVERAGE(J2:J13)"
    sn.cell(row=16, column=11).value = "=AVERAGE(K2:K13)"
    sn.cell(row=16, column=16).value = "=AVERAGE(P2:P13)"

    sn.cell(row=17, column=3).value = "=STDEV(C2:C13)"
    sn.cell(row=17, column=4).value = "=STDEV(D2:D13)"
    sn.cell(row=17, column=5).value = "=STDEV(E2:E13)"
    sn.cell(row=17, column=6).value = "=STDEV(F2:F13)"
    sn.cell(row=17, column=7).value = "=STDEV(G2:G13)"
    sn.cell(row=17, column=8).value = "=STDEV(H2:H13)"
    sn.cell(row=17, column=9).value = "=STDEV(I2:I13)"
    sn.cell(row=17, column=10).value = "=STDEV(J2:J13)"
    sn.cell(row=17, column=11).value = "=STDEV(K2:K13)"
    sn.cell(row=17, column=16).value = "=STDEV(P2:P13)"

    sn.cell(row=18, column=3).value = "=SUM(C17/C16)"
    sn.cell(row=18, column=4).value = "=SUM(D17/D16)"
    sn.cell(row=18, column=5).value = "=SUM(E17/E16)"
    sn.cell(row=18, column=6).value = "=SUM(F17/F16)"
    sn.cell(row=18, column=7).value = "=SUM(G17/G16)"
    sn.cell(row=18, column=8).value = "=SUM(H17/H16)"
    sn.cell(row=18, column=9).value = "=SUM(I17/I16)"
    sn.cell(row=18, column=10).value = "=SUM(J17/J16)"
    sn.cell(row=18, column=11).value = "=SUM(K17/K16)"
    sn.cell(row=18, column=16).value = "=SUM(P17/P16)"

    sn.cell(row=19, column=3).value = "=SUM(C2:C13)"
    sn.cell(row=19, column=4).value = "=SUM(D2:D13)"
    sn.cell(row=19, column=5).value = "=SUM(E2:E13)"
    sn.cell(row=19, column=6).value = "=SUM(F2:F13)"
    sn.cell(row=19, column=7).value = "=SUM(G2:G13)"
    sn.cell(row=19, column=8).value = "=SUM(H2:H13)"
    sn.cell(row=19, column=9).value = "=SUM(I2:I13)"
    sn.cell(row=19, column=10).value = "=SUM(J2:J13)"
    sn.cell(row=19, column=11).value = "=SUM(K2:K13)"
    sn.cell(row=19, column=16).value = "=SUM(P2:P13)"

    sn.cell(row=21, column=6).value = "=SUM(D16+F16)"
    sn.cell(row=22, column=6).value = "=SUM(E16+G16)"
    sn.cell(row=23, column=6).value = "=SUM(H16*1)"
    sn.cell(row=24, column=6).value = "=SUM(I16*1)"
    sn.cell(row=25, column=6).value = "=SUM(J16*1)"
    sn.cell(row=26, column=6).value = "=SUM(P16*1)"

C = 0
for sht in Teams:
    WS1.cell(row=2 + C, column=1).value = (sht[1:])
    C += 1
Stuff = ['Team #', 'Average Upper', 'Average Lower', 'Rotation', 'Position', 'Climb', 'Av Point Contrib']
s = 0
for st in Stuff:
    WS1.cell(row=1, column=1 + s).value = str(Stuff[s])
    s += 1

Tnmt.save("EagleRunner/Spreadsheets/Temp.xlsx")

Trnmt = openpyxl.load_workbook("EagleRunner/Spreadsheets/Temp.xlsx", read_only=False, keep_vba=True, data_only=False)

SN = Trnmt["Important Stuff"]

D = 0
for tn in Teams:
    TT = "=" + tn + "!" + "F21"
    TU = "=" + tn + "!" + "F22"
    TV = "=" + tn + "!" + "F23"
    TW = "=" + tn + "!" + "F24"
    TX = "=" + tn + "!" + "F25"
    TY = "=" + tn + "!" + "P16"
    SN.cell(row=2 + D, column=2).value = TT
    SN.cell(row=2 + D, column=3).value = TU
    SN.cell(row=2 + D, column=4).value = TV
    SN.cell(row=2 + D, column=5).value = TW
    SN.cell(row=2 + D, column=6).value = TX
    SN.cell(row=2 + D, column=7).value = TY
    SN.cell(row=2 + D, column=2).number_format = '0.00'
    SN.cell(row=2 + D, column=3).number_format = '0.00'
    SN.cell(row=2 + D, column=4).number_format = '0.00'
    SN.cell(row=2 + D, column=5).number_format = '0.00'
    SN.cell(row=2 + D, column=6).number_format = '0.00'
    SN.cell(row=2 + D, column=7).number_format = '0.00'

    D += 1

SN.conditional_formatting.add('B2:B75',
                              CellIsRule(operator='greaterThan', formula=[20.1], stopIfTrue=True, fill=greenFill))
SN.conditional_formatting.add('B2:B75',
                              CellIsRule(operator='between', formula=[10.1, 20.0], stopIfTrue=True, fill=yellowFill))
SN.conditional_formatting.add('B2:B75',
                              CellIsRule(operator='between', formula=[.01, 10], stopIfTrue=True, fill=redFill))
SN.conditional_formatting.add('C2:C75',
                              CellIsRule(operator='greaterThan', formula=[10.00], stopIfTrue=True, fill=greenFill))
SN.conditional_formatting.add('C2:C75',
                              CellIsRule(operator='between', formula=[5.1, 9.99], stopIfTrue=True, fill=yellowFill))
SN.conditional_formatting.add('C2:C75',
                              CellIsRule(operator='between', formula=[.01, 5], stopIfTrue=True, fill=redFill))
SN.conditional_formatting.add('D2:D75',
                              CellIsRule(operator='greaterThan', formula=[.75], stopIfTrue=True, fill=greenFill))
SN.conditional_formatting.add('D2:D75',
                              CellIsRule(operator='between', formula=[.5, .74], stopIfTrue=True, fill=yellowFill))
SN.conditional_formatting.add('D2:D75',
                              CellIsRule(operator='between', formula=[.01, .49], stopIfTrue=True, fill=redFill))
SN.conditional_formatting.add('E2:E75',
                              CellIsRule(operator='greaterThan', formula=[.75], stopIfTrue=True, fill=greenFill))
SN.conditional_formatting.add('E2:E75',
                              CellIsRule(operator='between', formula=[.5, .74], stopIfTrue=True, fill=yellowFill))
SN.conditional_formatting.add('E2:E75',
                              CellIsRule(operator='between', formula=[.01, .49], stopIfTrue=True, fill=redFill))
SN.conditional_formatting.add('F2:F75',
                              CellIsRule(operator='greaterThan', formula=[.75], stopIfTrue=True, fill=greenFill))
SN.conditional_formatting.add('F2:F75',
                              CellIsRule(operator='between', formula=[.5, .74], stopIfTrue=True, fill=yellowFill))
SN.conditional_formatting.add('F2:F75',
                              CellIsRule(operator='between', formula=[.01, .49], stopIfTrue=True, fill=redFill))
SN.conditional_formatting.add('G2:G75',
                              CellIsRule(operator='greaterThan', formula=[120], stopIfTrue=True, fill=greenFill))
SN.conditional_formatting.add('G2:G75',
                              CellIsRule(operator='between', formula=[75, 119.9], stopIfTrue=True, fill=yellowFill))
SN.conditional_formatting.add('G2:G75',
                              CellIsRule(operator='between', formula=[.01, 74.9], stopIfTrue=True, fill=redFill))
Trnmt.save("Tournaments/All_Tournaments2021.xlsx")

book = openpyxl.load_workbook("Tournaments/All_Tournaments2021.xlsx")
writer = pd.ExcelWriter("Tournaments/All_Tournaments2021.xlsx", engine='openpyxl')
writer.book = book

df1 = pd.read_excel("Match_Schedule/Match Schedule.xlsx", sheet_name='Match Schedule',
                    header=0)
df2 = pd.read_excel("Match_Schedule/Match Schedule.xlsx", sheet_name=str(TEAM) + ' Schedule',
                    header=0)
df3 = pd.read_excel("Tournaments/All_Tournaments2021.xlsx", sheet_name='Predictions')
df1.to_excel(writer, sheet_name='Match Schedule', index=False)
df2.to_excel(writer, sheet_name=str(TEAM) + ' Schedule', index=False)

prd = book["Predictions"]
prd.cell(row=1, column=1).value = 'Match #'
prd.cell(row=1, column=2).value = 'Red'
prd.cell(row=1, column=3).value = 'Blue'

rows = len(df2.index)
for r in range(0, rows):
    m = df2.at[r, 'Match #']
    prd.cell(row=r + 2, column=1).value = m

for m in range(0, rows):
    R1 = "=T" + str(df2.at[0 + int(m), 'Red 1']) + "!P16"
    R2 = "=T" + str(df2.at[0 + int(m), 'Red 2']) + "!P16"
    R3 = "=T" + str(df2.at[0 + int(m), 'Red 3']) + "!P16"
    prd.cell(row=100 + (m * 3), column=5).value = R1
    prd.cell(row=101 + (m * 3), column=5).value = R2
    prd.cell(row=102 + (m * 3), column=5).value = R3
    prd.cell(row=2 + m, column=2).value = "=SUM(E" + str(100 + (m * 3)) + ":E" + str(102 + (m * 3))
    B1 = "=T" + str(df2.at[0 + int(m), 'Blue 1']) + "!P16"
    B2 = "=T" + str(df2.at[0 + int(m), 'Blue 2']) + "!P16"
    B3 = "=T" + str(df2.at[0 + int(m), 'Blue 3']) + "!P16"
    prd.cell(row=100 + (m * 3), column=6).value = B1
    prd.cell(row=101 + (m * 3), column=6).value = B2
    prd.cell(row=102 + (m * 3), column=6).value = B3
    prd.cell(row=2 + m, column=3).value = "=SUM(F" + str(100 + (m * 3)) + ":F" + str(102 + (m * 3))

writer.save()
writer.close()

MS = pd.read_excel("Match_Schedule/Match Schedule.xlsx", sheet_name='Match Schedule')
MS.to_csv("Match_Schedule/Match Schedule.csv",
          columns=('Match #', 'Time', 'Red 1', 'Red 2', 'Red 3', 'Blue 1', 'Blue 2', 'Blue 3'), index=False)
