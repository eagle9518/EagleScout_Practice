# import os
import tbapy
from openpyxl import Workbook

YEAR = '2020'

Sched = Workbook()
ws = Sched.active
ws.title = "All_Tournaments"

tba = tbapy.TBA('opXlAfkuD4tQbDm2iskpBHdyYQbarWsQoeSG8w6MSKQ0c8jtbOnbREQu7z7nfUCK')

schedule_keys = tba.events(YEAR, simple=False, keys=False)
for i in range(1, len(schedule_keys)):
    cd = schedule_keys[i].event_code
    nm = schedule_keys[i].name
    sd = schedule_keys[i].start_date
    tp = schedule_keys[i].event_type_string
    ws.cell(row=1 + i, column=1).value = cd
    ws.cell(row=1 + i, column=2).value = nm
    ws.cell(row=1 + i, column=3).value = sd
    ws.cell(row=1 + i, column=4).value = tp

Sched.save('Tournaments\All_Tournaments' + YEAR + '.xlsx')
