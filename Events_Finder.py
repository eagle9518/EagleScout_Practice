import tbapy
from openpyxl import Workbook

YEAR = '2020'

schedule = Workbook()
work_sheet = schedule.active
work_sheet.title = 'Scheduled Tournaments'

tba = tbapy.TBA('opXlAfkuD4tQbDm2iskpBHdyYQbarWsQoeSG8w6MSKQ0c8jtbOnbREQu7z7nfUCK')

schedule_keys = tba.events(YEAR, simple=False, keys=False)
for i in range(len(schedule_keys)):
    ec = schedule_keys[i].event_code
    nm = schedule_keys[i].name
    sd = schedule_keys[i].start_date
    et = schedule_keys[i].event_type

    work_sheet.cell(row=1 + i, column=1).value = ec
    work_sheet.cell(row=1 + i, column=2).value = nm
    work_sheet.cell(row=1 + i, column=3).value = sd
    work_sheet.cell(row=1 + 1, column=4).value = et

Schedule = schedule.save(filename='Tournaments/All_Tournaments.xlsx')
